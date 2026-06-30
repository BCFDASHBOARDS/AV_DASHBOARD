# ============================================================
#  generate_dashboard.py
#  Liest WB_Übersicht.xlsx und schreibt dashboard.html neu.
#  Nur der const-DATA-Block wird ersetzt – das Design bleibt.
# ============================================================

import json, re, sys
from pathlib import Path
from datetime import date

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl fehlt. Bitte installieren: pip install openpyxl")

XLSX   = Path(r"C:\Users\wibau\OneDrive - Baussmann\2026\WB_Übersicht.xlsx")
DASH   = Path(__file__).parent / "dashboard.html"

def num(v):
    try:    return float(v) if v else 0.0
    except: return 0.0

def t(kg):
    return round(kg / 1000, 2)

# ---- Excel lesen ----
wb = openpyxl.load_workbook(str(XLSX), data_only=True)
ws = wb["Verbrauch N YTD"]

# Aggregat-Zeilen einlesen (nr | name → offeneP)
agg = {}
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    nr, name = row[1], row[2]
    if nr is None or name is None:
        continue
    key = (str(nr).strip(), str(name).strip())
    agg[key] = num(row[3])   # Spalte D = Offene P

def kg(nr, name):
    return agg.get((nr, name), 0.0)

# ---- DATA-Struktur aufbauen ----
data = {
    "stand":   str(date.today()),
    "einheit": "t",
    "quelle":  "WB_Übersicht.xlsx · Tab 'Verbrauch N YTD' · Spalte D (Offene P)",
    "bereiche": [
        {
            "name":    "Presserei",
            "summe_t": t(kg("Presserei", "Gesamt ")),   # trailing space wie in Excel
            "gruppen": [
                {
                    "name":    "Nagelpressen",
                    "summe_t": t(kg("Nagelpressen", "Gesamt")),
                    "cluster": [
                        {"name": "D9-1",               "wert_t": t(kg("Nagelpressen", "D9-1"))},
                        {"name": "Stiftdraht verzinkt", "wert_t": t(kg("Nagelpressen", "Stiftdraht verzinkt"))},
                        {"name": "C60 (Pins)",          "wert_t": t(kg("Nagelpressen", "C60 (Pins)"))},
                        {"name": "V2A 1.4307",          "wert_t": t(kg("Nagelpressen", "V2A 1.4307"))},
                        {"name": "V4A 1.4401",          "wert_t": t(kg("Nagelpressen", "V4A 1.4401"))},
                    ],
                },
                {
                    "name":    "Doppeldruck",
                    "summe_t": t(kg("Doppeldruck", "Gesamt")),
                    "cluster": [
                        {"name": "20MnB4",     "wert_t": t(kg("Doppeldruck", "20MnB4"))},
                        {"name": "1.4567 V2A", "wert_t": t(kg("Doppeldruck", "1.4567 V2A"))},
                        {"name": "1.4578 V4A", "wert_t": t(kg("Doppeldruck ", "1.4578 V4A "))},  # trailing spaces
                        {"name": "C60 (BSN)",  "wert_t": t(kg("Doppeldruck", "C60 (BSN)"))},
                    ],
                },
            ],
        },
        {
            "name":    "Klammern",
            "summe_t": t(kg("Klammern", "Gesamt")),
            "gruppen": [
                {
                    "name":    "Eindraht",
                    "summe_t": t(kg("Eindraht", "Gesamt")),
                    "cluster": [
                        {"name": "ED Verzinkt", "wert_t": t(kg("ED Verzinkt", "Gesamt"))},
                        {"name": "ED VA",       "wert_t": t(kg("ED VA",       "Gesamt"))},
                    ],
                },
                {
                    "name":    "Band",
                    "summe_t": t(kg("Band", "Gesamt")),
                    "cluster": [
                        {"name": "Band gesamt", "wert_t": t(kg("Band", "Gesamt"))},
                    ],
                },
            ],
        },
    ],
    "gesamt_t": t(kg("Presserei", "Gesamt ") + kg("Klammern", "Gesamt")),
}

# ---- dashboard.html patchen ----
html = DASH.read_text(encoding="utf-8")
new_line = f"const DATA = {json.dumps(data, ensure_ascii=False)};"
html_new = re.sub(r"const DATA = \{.*?\};", new_line, html, flags=re.DOTALL)

DASH.write_text(html_new, encoding="utf-8")
print(f"OK  gesamt={data['gesamt_t']} t  Presserei={data['bereiche'][0]['summe_t']} t  "
      f"Klammern={data['bereiche'][1]['summe_t']} t  Stand={data['stand']}")
