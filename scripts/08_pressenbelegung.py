# ============================================================
# 08_pressenbelegung.py
# Vollstaendige Pressenbelegung-Extraktion:
#   - Maschinenauftraege aus allen Tabs
#   - NQ03-013: nur -200-Suffix zählt (Sonderlinie, kein Doppelzählen)
#   - Unverteilte FA aus "Datenimport PW" (Status=nein)
#     → wahrscheinliche Maschinenzuordnung per Durchmesser/Laenge/Suffix
#     → kg aus "Gewichte"-Tab, Stunden per Planleistung
#   - kg-Plausifilter: >500 kg/‰ → verwerfen
# Ausgabe: _data/pressenbelegung.json + docs/_data/pressenbelegung.json
# ============================================================

import openpyxl, json, os, re, sys
from datetime import datetime

# Pfade relativ zu diesem Skript
_HERE = os.path.dirname(os.path.abspath(__file__))
_BASE = os.path.dirname(_HERE)
XLS      = os.path.join(_BASE, "_source", "Planung_Pressen_NEU.xlsx")
OUT_DATA = os.path.join(_BASE, "_data",   "pressenbelegung.json")
OUT_DOCS = os.path.join(_BASE, "docs", "_data", "pressenbelegung.json")

KG_PER_MILLE_MAX = 500  # max plausibles kg/‰ Verhältnis

# ---- Mappings -------------------------------------------------------
MASCHINE_GRUPPE = {
    "N51-2":"DKopf_Offset","N90-D2":"DKopf_Offset","N90-D3":"DKopf_Offset",
    "N90-D4":"DKopf_Offset","NQ03-013":"DKopf_Offset",
    "N31-1":"N31","N31-2":"N31","N31-3":"N31",
    "N41-7":"N41","N41-6":"N41","N41-2":"N41","N41-5":"N41","N41-1":"N41","N41-3":"N41",
    "N51-3":"N51","N51-1":"N51","N61-3":"N61","N61-2":"N61","N61-1":"N61",
    "N90-2,5":"N90","N90-3,1":"N90","N90-D1":"N90",
    "Enkotec 2,5 NA03":"Enkotec","Enkotec 2,8 NI01":"Enkotec","Enkotec 2,8 NU03":"Enkotec",
    "Hilgeland COH":"Doppeldruck","Hilgeland COLH":"Doppeldruck",
    "FWB 20C-1 weiß":"Doppeldruck","FWB 20C-2 grün":"Doppeldruck",
    "Hilgeland HD7":"Doppeldruck","Hilgeland HD6-40":"Doppeldruck",
    "Hilgeland HD6-60":"Doppeldruck","Hilgeland HC5-60":"Doppeldruck",
    "Hilgeland C2AZ":"Doppeldruck","ChunZu CH12LL":"Doppeldruck","Klose MTH350":"Doppeldruck",
}
SCHICHTEN = {"DKopf_Offset":2,"N31":2,"N41":2,"N51":2,"N61":2,"N90":2,"Enkotec":2,"Doppeldruck":1}
GRUPPEN_ORDER = ["DKopf_Offset","N31","N41","N51","N61","N90","Enkotec","Doppeldruck"]
GRUPPEN_LABEL = {
    "DKopf_Offset":"D-Kopf + Offset","N31":"N31","N41":"N41","N51":"N51",
    "N61":"N61","N90":"N90","Enkotec":"Enkotec","Doppeldruck":"Doppeldruck"
}
# Spaltenstruktur je Quell-Tab: (col_folgeAG, col_bemerkung, col_stunden)
TAB_STD_COL = {
    "Doppeldruck": (8, None, 9),
    "N31+41":      (8, 9, 10),
    "N51+61":      (8, 9, 10),
    "N90 + ENK":   (8, 9, 10),
    "NQ03-013":    (8, 10, 11),
}

def normalize(s):
    return ' '.join(str(s).split()).strip('\xa0').strip()

def get_zustand(s):
    s = str(s).lower()
    if 'freigegeb' in s: return 'freigegeben'
    if 'terminiert' in s: return 'terminiert'
    if 'unterbrochen' in s: return 'unterbrochen'
    return 'unbekannt'

def get_suffix(artnr):
    m = re.search(r'-(\d{3})$', artnr)
    return m.group(1) if m else ''

def parse_artnr_dims(artnr):
    """Gibt (diam_mm, length_mm, suffix_str) zurück oder (None,None,None)."""
    mo = re.match(r'^\d{2}-(\d{6})-(\d{3})$', artnr)
    if mo:
        mid = mo.group(1)
        suf = mo.group(2)
        return int(mid[:3]) / 100.0, int(mid[3:]), suf
    return None, None, None

def assign_machine(artnr, typ, diam, length, suffix, loads):
    """Gibt (machine_name, group_key) zurück."""
    is_d_kopf = suffix.startswith('2') and suffix not in ('001',)
    is_offset  = suffix.startswith('3') and suffix not in ('001',)

    if typ == 'DD':
        dd_specs = [
            ('Hilgeland COH',   3.1,  45),
            ('Hilgeland COLH',  3.1,  65),
            ('FWB 20C-1 weiß',  3.1,  80),
            ('FWB 20C-2 grün',  3.1,  80),
            ('Hilgeland HC5-60',5.0,  65),
            ('Hilgeland HD6-40',6.0,  40),
            ('Hilgeland HD6-60',6.0,  60),
            ('Hilgeland HD7',   7.0, 170),
            ('Hilgeland C2AZ',  8.0, 100),
            ('ChunZu CH12LL',  12.7, 203),
            ('Klose MTH350',   10.0, 350),
        ]
        cands = [(m, md, ml) for m, md, ml in dd_specs if diam <= md and length <= ml]
        if not cands:
            return 'Hilgeland C2AZ', 'Doppeldruck'
        cands.sort(key=lambda x: x[1] * x[2])
        return cands[0][0], 'Doppeldruck'

    # NP: D-Kopf (-2xx)
    if is_d_kopf:
        if suffix == '200' and abs(diam - 2.8) < 0.05 and 50 <= length <= 80:
            return 'NQ03-013', 'DKopf_Offset'
        elif diam >= 2.4 and length >= 48:
            cands = ['N90-D2', 'N90-D3', 'N90-D4']
            return min(cands, key=lambda m: loads.get(m, 0)), 'DKopf_Offset'
        else:
            return 'N51-2', 'DKopf_Offset'

    # NP: Offset (-3xx)
    if is_offset:
        return 'N51-2', 'DKopf_Offset'

    # Standard: nach Durchmesser + Länge
    if diam >= 3.4:
        cands = ['N61-3', 'N61-2', 'N61-1']
        return min(cands, key=lambda m: loads.get(m, 0)), 'N61'

    if 2.4 <= diam <= 3.8 and 48 <= length <= 100:
        cands = ['N90-2,5', 'N90-3,1', 'N90-D1']
        return min(cands, key=lambda m: loads.get(m, 0)), 'N90'

    if 2.2 <= diam <= 4.2 and length > 80:
        cands = ['N51-3', 'N51-1']
        return min(cands, key=lambda m: loads.get(m, 0)), 'N51'

    if 1.8 <= diam <= 3.4 and length <= 80:
        cands = ['N41-7', 'N41-6', 'N41-2', 'N41-5', 'N41-1', 'N41-3']
        return min(cands, key=lambda m: loads.get(m, 0)), 'N41'

    if diam <= 2.3:
        n31_max = {'N31-1': 55, 'N31-2': 50, 'N31-3': 48}
        elig = [m for m, ml in n31_max.items() if length <= ml]
        if elig:
            return min(elig, key=lambda m: loads.get(m, 0)), 'N31'

    # Fallback
    cands = ['N41-7', 'N41-6', 'N41-2', 'N41-5', 'N41-1', 'N41-3']
    return min(cands, key=lambda m: loads.get(m, 0)), 'N41'

# =====================================================================
# MAIN
# =====================================================================
if not os.path.exists(XLS):
    print(f"[ERR] Quelldatei nicht gefunden: {XLS}", file=sys.stderr)
    sys.exit(1)

print("=== 08_pressenbelegung.py ===")
wb = openpyxl.load_workbook(XLS, read_only=True, data_only=True)

# 1. Gewichte-Lookup --------------------------------------------------
gewichte = {}
ws_g = wb["Gewichte"]
for row in ws_g.iter_rows(min_row=2, values_only=True):
    if not row[0]: break
    gewichte[str(row[0]).strip()] = row[2]
print(f"  Gewichte geladen: {len(gewichte)} Artikel")

# 2. Maschinenköpfe + bestehende Aufträge ------------------------------
all_maschinen = {}
machine_planl = {}  # name -> ‰/h
TABS = ["Doppeldruck", "N31+41", "N51+61", "N90 + ENK", "NQ03-013"]

for tab in TABS:
    ws = wb[tab]
    akt = None
    ci_folge, ci_bemerk, ci_std = TAB_STD_COL[tab]

    for row in ws.iter_rows(max_row=300, values_only=True):
        cols = list(row) + [None] * 15
        c0 = str(cols[0]).strip() if cols[0] is not None else ''
        c1 = cols[1]
        c5 = str(cols[5]).strip() if cols[5] is not None else ''

        is_mach = (c5 == '‰' and isinstance(cols[4], (int, float))) or \
                  (str(c1).strip() == 'NQ03-013' and cols[0] is None)
        if is_mach:
            akt = normalize(str(c1 if cols[0] is None else cols[0]))
            if isinstance(cols[8], (int, float)) and cols[8] > 0:
                machine_planl[akt] = float(cols[8])
            if akt not in all_maschinen:
                all_maschinen[akt] = {'name': akt, 'auftraege': [], 'sum_pm': 0, 'sum_kg': 0.0, 'sum_h': 0.0}
            continue

        if c0 == 'S' and str(c1).strip() == 'FA':
            continue

        if akt and isinstance(c1, (int, float)) and c1 and float(c1) > 0:
            skip = {'nicht vorhanden', '#n/a', '#wert!', 'fehlt', ''}
            if c0.lower() in skip or c0.startswith('#'): continue
            artnr = str(cols[2]).strip() if cols[2] else ''
            if not artnr or artnr.lower() in skip: continue

            suffix = get_suffix(artnr)
            # NQ03-013: nur -200 Suffix aufnehmen (Sonderlinie — -202/-207 sind Zwischenstufen)
            is_nq = (akt == 'NQ03-013')
            if is_nq and suffix != '200':
                continue

            menge = round(float(cols[4])) if isinstance(cols[4], (int, float)) else None
            ist   = round(float(cols[5])) if isinstance(cols[5], (int, float)) else None
            kg_raw = cols[6]
            kg = None
            if isinstance(kg_raw, (int, float)) and kg_raw > 0:
                ratio = float(kg_raw) / menge if menge and menge > 0 else 9999
                if ratio > KG_PER_MILLE_MAX:
                    print(f"  FILTER {akt} FA {int(float(c1))} {artnr}: {kg_raw:,.0f}kg/{menge}‰")
                else:
                    kg = round(float(kg_raw), 1)

            folge  = str(cols[ci_folge]).strip() if cols[ci_folge] else ''
            bemerk = str(cols[ci_bemerk]).strip() if ci_bemerk and cols[ci_bemerk] else ''
            std    = round(float(cols[ci_std]), 1) if isinstance(cols[ci_std], (int, float)) else None
            termin = cols[7].strftime('%Y-%m-%d') if hasattr(cols[7], 'strftime') else None
            kurztext = str(cols[3]).strip() if cols[3] else ''

            m = all_maschinen[akt]
            m['auftraege'].append({
                'zustand': get_zustand(c0), 'fa': int(float(c1)), 'artnr': artnr,
                'kurztext': kurztext, 'menge': menge, 'ist': ist, 'kg': kg, 'termin': termin,
                'folge_ag': folge, 'bemerkung': bemerk, 'stunden': std
            })
            m['sum_pm'] += menge or 0
            if kg: m['sum_kg'] += kg
            if std: m['sum_h'] += std

print(f"  Maschinen extrahiert: {len(all_maschinen)}")

# 3. Implied Planleistung für DD + NQ ---------------------------------
for mname, m in all_maschinen.items():
    if mname not in machine_planl and m['sum_h'] > 0 and m['sum_pm'] > 0:
        machine_planl[mname] = m['sum_pm'] / m['sum_h']

# 4. Unverteilte FAs (Status=nein) ------------------------------------
ws_pw = wb["Datenimport PW"]
unassigned = []
for row in ws_pw.iter_rows(min_row=2, max_row=600, values_only=True):
    if row[0] is None: break
    if str(row[7]).lower().strip() == 'nein':
        unassigned.append({
            'fa': row[0], 'zustand': str(row[1]), 'artnr': str(row[2]).strip(),
            'kurztext': str(row[3]).strip(), 'typ': str(row[4]).strip(),
            'menge': int(row[5]) if row[5] else 0, 'nachfolge': str(row[6] or '')
        })
print(f"  Unverteilte FAs: {len(unassigned)}")

# 5. Zuordnung + Extrapolation ----------------------------------------
machine_loads = {mname: m['sum_h'] for mname, m in all_maschinen.items()}
extrapol_count = 0
extrapol_pm = 0
extrapol_kg = 0.0
skipped = 0

for fa in unassigned:
    artnr = fa['artnr']
    diam, length, suffix = parse_artnr_dims(artnr)
    if diam is None:
        print(f"  [SKIP] FA {fa['fa']} {artnr}: Artnr nicht parsbar")
        skipped += 1
        continue

    machine, gkey = assign_machine(artnr, fa['typ'], diam, length, suffix, machine_loads)

    menge = fa['menge']
    ref_wt = gewichte.get(artnr)
    kg = round(menge * ref_wt, 1) if ref_wt and menge else None
    if kg is not None and menge and (kg / menge) > KG_PER_MILLE_MAX:
        print(f"  KG-FILTER extr. FA {fa['fa']}: {kg}kg/{menge}‰")
        kg = None

    planl = machine_planl.get(machine)
    hours = round(menge / planl, 1) if planl and planl > 0 and menge else None

    if machine not in all_maschinen:
        all_maschinen[machine] = {'name': machine, 'auftraege': [], 'sum_pm': 0, 'sum_kg': 0.0, 'sum_h': 0.0}

    m = all_maschinen[machine]
    m['sum_pm'] += menge
    if kg:    m['sum_kg'] += kg
    if hours: m['sum_h']  += hours
    m['auftraege'].append({
        'zustand': 'geplant', 'fa': fa['fa'], 'artnr': artnr, 'kurztext': fa['kurztext'],
        'menge': menge, 'ist': None, 'kg': kg, 'termin': None,
        'folge_ag': fa['nachfolge'], 'bemerkung': '', 'stunden': hours, 'extrapoliert': True
    })
    machine_loads[machine] = m['sum_h']
    extrapol_count += 1
    extrapol_pm += menge
    if kg: extrapol_kg += kg

print(f"  Extrapoliert: {extrapol_count} FA, {extrapol_pm:,} ‰, {extrapol_kg:,.0f} kg")
if skipped:
    print(f"  Uebersprungen: {skipped} FA (Artnr nicht parsbar)")

# 6. Gruppen + JSON ---------------------------------------------------
gruppen = {k: {'key': k, 'name': GRUPPEN_LABEL[k], 'permille': 0, 'kg': 0.0, 'stunden': 0.0, 'maschinen': []}
           for k in GRUPPEN_ORDER}

for mname, m in all_maschinen.items():
    norm = normalize(mname)
    gkey = MASCHINE_GRUPPE.get(norm)
    if not gkey:
        nl = norm.lower()
        if nl.startswith('n31'): gkey = 'N31'
        elif nl.startswith('n41'): gkey = 'N41'
        elif nl.startswith('n51'): gkey = 'N51'
        elif nl.startswith('n61'): gkey = 'N61'
        elif nl.startswith('n90'): gkey = 'N90'
        elif nl.startswith('enkotec'): gkey = 'Enkotec'
        elif any(nl.startswith(p) for p in ['hilgeland', 'fwb', 'chunzu', 'klose']): gkey = 'Doppeldruck'
        else:
            print(f"  [WARN] Keine Gruppe: {mname}")
            continue

    sch = SCHICHTEN[gkey]
    sum_h  = round(m['sum_h'], 1)
    sum_pm = round(m['sum_pm'])
    sum_kg = round(m['sum_kg'], 1)
    auslast = round(sum_h / (sch * 8.0), 2) if sum_h > 0 else 0.0

    g = gruppen[gkey]
    g['maschinen'].append({
        'name': m['name'], 'permille': sum_pm, 'kg': sum_kg, 'stunden': sum_h,
        'auslastung_tage': auslast, 'schichten': sch, 'auftraege': m['auftraege']
    })
    g['permille'] += sum_pm
    g['kg']       += sum_kg
    g['stunden']  += sum_h

gruppen_list = []
for gkey in GRUPPEN_ORDER:
    g = gruppen[gkey]
    g['permille'] = round(g['permille'])
    g['kg']       = round(g['kg'], 1)
    g['stunden']  = round(g['stunden'], 1)
    gruppen_list.append(g)

kpi = {
    'gesamt_permille': sum(g['permille'] for g in gruppen_list),
    'gesamt_kg':       round(sum(g['kg'] for g in gruppen_list), 1),
    'gesamt_stunden':  round(sum(g['stunden'] for g in gruppen_list), 1)
}
meta = {
    'extrapoliert_count':    extrapol_count,
    'extrapoliert_permille': extrapol_pm,
    'extrapoliert_kg':       round(extrapol_kg, 1)
}
result = {'kpi': kpi, 'meta': meta, 'gruppen': gruppen_list, 'timestamp': datetime.now().isoformat()}

for out in [OUT_DATA, OUT_DOCS]:
    os.makedirs(os.path.dirname(out), exist_ok=True)
    with open(out, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

print(f"\n[OK] KPI: {kpi['gesamt_permille']:,} ‰  |  {kpi['gesamt_kg']:,.0f} kg ({kpi['gesamt_kg']/1000:.0f} t)  |  {kpi['gesamt_stunden']:,.0f} h")
print(f"     Prognose: +{extrapol_pm:,} ‰  +{extrapol_kg:,.0f} kg")
